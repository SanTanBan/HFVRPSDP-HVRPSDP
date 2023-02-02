#import time
import pandas as pd
import matplotlib.pylab as plt
from gurobipy import *
from winsound import Beep
import os
import openpyxl
import GeneRead

# Refer:- https://www.gurobi.com/documentation/9.5/refman/variables.html
# Gurobi has some Tolerances on the Binary and Integer Variables...
# For Tolerances:- https://www.gurobi.com/documentation/9.5/refman/intfeastol.html#parameter:IntFeasTol
tolerance=-1e-5 # This is the default value for Gurobi

def Avci_Exact_Algorithm(max_seconds_allowed_for_calculation=86400):

    directory_name="AvciExact"
    main_dir = directory_name
    os.mkdir(main_dir)
    directory_name=directory_name+"/"

    Depot_and_Relief_Centres,Latitude,Longitude,pp,d=GeneRead.Reader.Relief_Centre_Reader()
    Relief_Centres=Depot_and_Relief_Centres.copy()
    del Relief_Centres[0]
    Vehicle_Types,VN,VQ,VS,VC,Vehicle_Width=GeneRead.Reader.Vehicle_Type_Reader()


    set_of_Vehicles=[] # Type and Number
    for k in Vehicle_Types:
        for kk in range(int(VN[k])):
            set_of_Vehicles.append((k,kk))
    #print(set_of_Vehicles)

    #Creating the Distance Matrices
    C=GeneRead.Reader.Distance_Combinations_Reader()
    for i in Depot_and_Relief_Centres:
        for k in Vehicle_Types:
            C[(i,i,k)]=0

    x, y, z = {}, {}, {}
    # Set the problem
    mdl=Model("VRP")
    mdl.setParam('TimeLimit', max_seconds_allowed_for_calculation)
    

    # Decision Variables
    # Iff Arc joining i & j is included within the solution for the Layer k
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in set_of_Vehicles:
                x[i, j, k] = mdl.addVar(vtype=GRB.BINARY, name="x "+str(i)+" "+str(j)+" "+str(k))
                mdl.update()

    # Amount of collected load across Arc(i,j) by a Vehicle in Layer k
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in set_of_Vehicles:
                y[i, j, k] = mdl.addVar(vtype=GRB.CONTINUOUS, name="y "+str(i)+" "+str(j)+" "+str(k))
                mdl.update()

    # Amount of delivery load across Arc(i,j) done by a Vehicle in Layer k
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in set_of_Vehicles:
                z[i, j, k] = mdl.addVar(vtype=GRB.CONTINUOUS, name="z "+str(i)+" "+str(j)+" "+str(k))
                mdl.update()

    print("********** This Avci Topaloglu Formulation has updated Objective including a new subscript k in the Cost Matrix **********\n")

    """ Updated Objective includes a new subscript k """ # Set Objective Function (EQ 1) 
    mdl.setObjective(quicksum(VC[k[0]]*x[0,j,k] for k in set_of_Vehicles for j in Relief_Centres) + quicksum(VS[k[0]]*x[i,j,k]*C[i,j,k[0]] for k in set_of_Vehicles for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres),GRB.MINIMIZE)
    mdl.update()

    #Ensuring a single vehicle caters to a Relief Center (EQ 2)
    mdl.addConstrs(quicksum(x[i,j,k] for i in Depot_and_Relief_Centres for k in set_of_Vehicles) ==1 for j in Relief_Centres)
    mdl.update()

    # Ensuring equal number of Incoming and Outgoing paths are available from all Nodes (EQ 3)
    mdl.addConstrs(quicksum(x[i,p,k] for i in Depot_and_Relief_Centres) - quicksum(x[p,j,k] for j in Depot_and_Relief_Centres)==0  for k in set_of_Vehicles for p in Depot_and_Relief_Centres)
    mdl.update()

    # Ensuring at most 1 outgoing paths is available for each Vehicle of the VN[k] vehicles at the Depot of respective Vehicle Type (EQ 4))
    mdl.addConstrs(quicksum(x[0,j,k] for j in Relief_Centres) <= 1 for k in set_of_Vehicles)
    mdl.update()

    '''Flow Limitation Constraints'''

    #Ensuring initial PickUp from Nodes is 0 (EQ 5)
    mdl.addConstrs(y[0,j,k] == 0 for j in Relief_Centres for k in set_of_Vehicles)
    mdl.update()

    #Ensuring final Delivery to Nodes is 0 (EQ 6)
    mdl.addConstrs(z[i,0,k] == 0 for i in Relief_Centres for k in set_of_Vehicles)
    mdl.update() 

    #Ensuring the PickUp constraints are satisfied (EQ 7)
    mdl.addConstrs(quicksum(y[i,j,k] for j in Depot_and_Relief_Centres for k in set_of_Vehicles) - quicksum(y[j,i,k] for j in Depot_and_Relief_Centres for k in set_of_Vehicles) == pp[i] for i in Relief_Centres)
    mdl.update()

    #Ensuring the Delivery constraints are satisfied (EQ 8)
    mdl.addConstrs(quicksum(z[j,i,k] for j in Depot_and_Relief_Centres for k in set_of_Vehicles) - quicksum(z[i,j,k] for j in Depot_and_Relief_Centres for k in set_of_Vehicles) == d[i] for i in Relief_Centres)
    mdl.update()


    '''Constraining the Sum of Flows to and from the Origin/Depot/Warehouse/NDRF_BASE'''
    # Redundant
    #(EQ 9) Ensuring sum of all PickUp Flow Variables to the Origin [0th Node] is equal to the total PickUps of all Nodes
    mdl.addConstr(quicksum(y[i,0,k] for i in Relief_Centres for k in set_of_Vehicles)  == quicksum(pp[i] for i in Relief_Centres))
    mdl.update()
    #(EQ 10) Ensuring sum of all Delivery Flow Variables from the Origin [0th Node] is equal to the total Demand of all Nodes
    mdl.addConstr(quicksum(z[(0,i,k)] for i in Relief_Centres for k in set_of_Vehicles) == quicksum(d[i] for i in Relief_Centres))
    mdl.update()


    # Ensuring the vehicle capacity is never exceeded (EQ 11)
    mdl.addConstrs(y[i,j,k]+z[i,j,k] <= VQ[k[0]]*x[i,j,k] for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in set_of_Vehicles)
    mdl.update()

    # Solve the Problem using default CBC
    #start_time=time.time()
    mdl.optimize()
    freq=500
    dur=1000
    Beep(freq, dur)

    OptimalityGap=mdl.MIPGap
    print("Final MIP gap value: %f" % OptimalityGap)
    best_bound=mdl.ObjBound
    print("The best Bound found is:- ",best_bound)
    Solutions_Found=mdl.SolCount
    objec_val=-1
    if Solutions_Found:
        objec_val=mdl.getObjective().getValue()

        #end_time=time.time()

        #winsound.Beep(555-19*upto_Node_number, 888+19*upto_Node_number) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
        #print("This is the status:- ", status)
        #print('Runtime is',mdl.Runtime)


        main_dir=main_dir+"/"
        # Draw the optimal routes Layerwise
        for k in set_of_Vehicles:
            plt.figure(figsize=(9,9))
            for i in Depot_and_Relief_Centres:
                if i==0:
                    plt.scatter(Longitude[i],Latitude[i], c='r',marker='s')
                    plt.text( Longitude[i] + 0.33, Latitude[i] + 0.33, "Depot")
                else:
                    plt.scatter(Longitude[i],Latitude[i], c='black')
                    plt.text( Longitude[i] + 0.33, Latitude[i] + 0.33, i)
            plt.title('mVRPSDC Tours for Vehicles of Type '+str(k)+" on the corresponding layer "+str(k))
            plt.ylabel("Latitude")
            plt.xlabel("Longitude")

            max=0   # Finding the maximum utilised vehicle capacity
            routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if ((x[i,j,k].x<=1-tolerance) and (x[i,j,k].x>=1+tolerance))]
            #routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j and p.value(x[i,j,k])==1]
            arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor='blue')
            for i, j in routes:
                utilized_capacity=y[i,j,k].x+z[i,j,k].x
                if utilized_capacity>max:
                    max=utilized_capacity
                plt.annotate('', xy=[Longitude[j], Latitude[j]], xytext=[Longitude[i], Latitude[i]], arrowprops=arrowprops)    
                #plt.text((Nodes.iloc[i]["Longitude"]+Nodes.iloc[j]["Longitude"])/2, (Nodes.iloc[i]["Latitude"]+Nodes.iloc[j]["Latitude"])/2, f'{utilized_capacity}',fontweight="bold")

            print("The maximum vehicle capacity utilised ever in any tour in layer (i.e. Vehicle) ",k," is: ",max," out of the total available",VQ[k[0]])
            
            
            used_vehicles=0 # Finding the maximum number of vehicles being used
            for j in Relief_Centres:
                used_vehicles=x[0,j,k].x+used_vehicles
            #print("The maximum numbers of vehicles used is: ",used_vehicles," out of total available ",VN[k])
            name="Vehicle "+str(k)+" and Capacity_ "+str(max)+"--"+str(VQ[k[0]])+" with Objective Value_ "+str(objec_val)+" & Solver Time is_ "+str(mdl.Runtime)+"seconds.eps"
            main_dir_for_Image=main_dir+"{}"
            plt.savefig(main_dir_for_Image.format(name),format='eps')

        # Writing the Routes in a Text File
        textfile = open(main_dir+"Vehicle Routes.txt","w")
        textfile.write("\t The Objective Value is "+str(objec_val)+" obtained within "+str(mdl.Runtime)+" seconds with Relative Gap of "+str(OptimalityGap*100)+"% \n \n \n")
        for k in set_of_Vehicles:
            for j in Relief_Centres:
                if (x[0,j,k].x<=1-tolerance) and (x[0,j,k].x>=1+tolerance):
                    start_node=j
                    textfile.write("Vehicle : "+str(k)+",\t Route=\t 0")
                    while start_node!=0:
                        textfile.write(" --> "+str(start_node))
                        for i in Depot_and_Relief_Centres:
                            if  start_node!=i and ((x[start_node,i,k].x<=1-tolerance) and (x[start_node,i,k].x>=1+tolerance)):
                                start_node=i
                                break
                    if start_node==0:
                        textfile.write(" --> "+str(start_node)+"\n")
        textfile.close()
        

        # Call a Workbook() function of openpyxl to create a new blank Workbook object
        wb_individual = openpyxl.Workbook()
        # Get workbook active sheet from the active attribute
        sheet_individual = wb_individual.active
        row_number_on_Individual_Sheet=1
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
        cell.value = "From Node i"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
        cell.value = "To Node j"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
        cell.value = "Vehicle k"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
        cell.value = "x_ijk indicating whether the Arc is selected"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
        cell.value = "y_ijk indicating the amount of Pickup"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
        cell.value = "z_ijk indicating the amount of Delivery"
        for i in Depot_and_Relief_Centres:
            for j in Depot_and_Relief_Centres:
                for k in set_of_Vehicles:
                    row_number_on_Individual_Sheet+=1
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
                    cell.value = i
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
                    cell.value = j
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
                    cell.value = str(k)
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
                    cell.value = x[i,j,k].x
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
                    cell.value = y[i,j,k].x
                    cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
                    cell.value = z[i,j,k].x
        wb_individual.save(str(main_dir)+"Solution Details.xlsx")



    return objec_val,mdl.Runtime,OptimalityGap,best_bound


# For more retrieving Solutions
# https://www.gurobi.com/documentation/9.5/refman/retrieving_solutions.html
# https://www.gurobi.com/documentation/9.5/refman/objbound.html#attr:ObjBound
# https://www.gurobi.com/documentation/9.5/refman/parameter_examples.html#sec:ParameterExamples