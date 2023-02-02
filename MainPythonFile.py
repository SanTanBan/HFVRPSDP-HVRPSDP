import openpyxl
import pandas as pd

location_of_Node_Locations_file="Node Specifications.xlsx"
destination_to_save_Distance_Matrix=""
Relief_Locations=pd.read_excel("Node Specifications.xlsx","Locations, PickUp & Delivery",index_col=0)

Vehicles=pd.read_excel("Vehicle Type Specifications.xlsx","Vehicle Specifications",index_col=0)
wb = openpyxl.Workbook()
for k,Vehicle_Specs in Vehicles.iterrows():

    #print(k)
    #print(Vehicle_Specs)
    ws = wb.create_sheet("Distances for VT "+str(k))
    row_number=1
    cell = ws.cell(row= row_number, column = 1)
    cell.value = "Origin Node"
    cell = ws.cell(row= row_number, column = 2)
    cell.value = "Destination Node"
    cell = ws.cell(row= row_number, column = 3)
    cell.value = "Distance"
    for i, row1 in Relief_Locations.iterrows():
        for j, row2 in Relief_Locations.iterrows():
                if i!=j:
                    row_number+=1
                    cell = ws.cell(row= row_number, column = 1)
                    cell.value = str(i)
                    cell = ws.cell(row= row_number, column = 2)
                    cell.value = str(j)
                    cell = ws.cell(row= row_number, column = 3)
                    cell.value = ((row1["Latitude"]-row2["Latitude"])**2+(row1["Longitude"]-row2["Longitude"])**2)**0.5
del wb["Sheet"]
wb.save(str(destination_to_save_Distance_Matrix)+"Network Distance Matrix.xlsx")

# After this the rest of the MainPythonFile starts:


# Dependency Installations Done (https://datatofish.com/command-prompt-python/)
# grbgetkey 7f5dd15a-3e33-11ed-9fdc-0242ac120004

# from GeneRead import Generator
# Generator.Node_Generator(555,latitude_uniform_distribution_upper_bound=125,pickup_quantity_uniform_distribution_upper_bound=125,longitude_uniform_distribution_upper_bound=125,delivery_quantity_uniform_distribution_upper_bound=125)
# Generator.Lp_Norm_Random_Matrix_for_each_Vehicle_Type_Generator(p=[1.8,1.6,1.4,1.2],Vehicle_Types=[1,2,3,4])


import winsound
from AvciTopalogluPaperFormulation import Avci_Exact_Algorithm
#from ATS_Algorithm import ATS_Algorithm
#from ATS_Flowchart import ATS_Flowchart
#from ATS_Text import ATS_Text
#from HLS_Algorithm import HLS_Algorithm
#from HLS_Flowchart import HLS_Flowchart
#from HLS_Text import HLS_Text
from SanBan_Algorithm import SANBAN_Gurobi_Algorithm
from SanBan_without_Redundants import SANBAN_withoutRedundant_Algorithm
from ATSAH import ATS_Algorithm_Heuristic
from ATSAML_6 import ATSAML_6
from ATSAML_10 import ATSAML_10
from ATSA_MT_19 import ATSA_MT_19
from ATSAML_6_JellyFish import ATSAMT_6_JF
from ATSAML_10_JellyFish import ATSAMT_10_JF
from ATSA_MT_19_JellyFish import ATSAMT_19_JF
from BarisKara import BarisKara_Gurobi_Algorithm
import cv2
#import openpyxl
from SanBan_FutureWorks import SANBAN_FutureWork
frameSize=(1920,1080)

Heuristic_Runs=30

# Call a Workbook() function of openpyxl to create a new blank Workbook object
Combined_Solution_Comparison= openpyxl.Workbook()
Solo_Sheet = Combined_Solution_Comparison.create_sheet("All for this problem size")

#Objective_Sheet = Combined_Solution_Comparison.create_sheet("Objectives")
#Time_Sheet = Combined_Solution_Comparison.create_sheet("Solution Time")
#Vehicle_Sheet = Combined_Solution_Comparison.create_sheet("Used Vehicles")
#Total_Iterations = Combined_Solution_Comparison.create_sheet("No. of Iterations")
#Per_Iteration_Span = Combined_Solution_Comparison.create_sheet("100 Iteration Span")
del Combined_Solution_Comparison["Sheet"]

# Providing Column Headings
cell = Solo_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = Solo_Sheet.cell(row = 1, column = 2)
cell.value = "Number of Nodes"
cell = Solo_Sheet.cell(row = 1, column = 3)
cell.value = "Vehicle Types Used"
cell = Solo_Sheet.cell(row = 1, column = 4)
cell.value = "Number of Vehicles available"


"""
num_of_run_ATSA=1
cell = Solo_Sheet.cell(row = 1, column = 10)
cell.value = "ATS Algorithm Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 11)
cell.value = "ATS Algorithm Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 12)
cell.value = "ATS Algorithm S.D. of Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 13)
cell.value = "ATS Algorithm Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 14)
cell.value = "ATS Algorithm S.D. of Solution Time"
min_objective=999999999999999999
average_objective=0
average_time=0
Objective_Array=[]
Time_Array=[]
for i in range(num_of_run_ATSA):
    x_b,Time=ATS_Algorithm(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    average_time+=Time
    Objective_Array.append(x_b[0])
    Time_Array.append(Time)
average_time/=num_of_run_ATSA
average_objective/=num_of_run_ATSA
SD_Objective=0
SD_Time=0
for i in Objective_Array:
    SD_Objective+=pow(i-average_objective,2)
SD_Objective/=num_of_run_ATSA
SD_Objective=pow(SD_Objective,0.5)
for i in Time_Array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSA
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 10)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 11)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 12)
cell.value = SD_Objective
cell = Solo_Sheet.cell(row = 2, column = 13)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 14)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


num_of_run_ATSF=1
cell = Solo_Sheet.cell(row = 1, column = 16)
cell.value = "ATS Flowchart Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 17)
cell.value = "ATS Flowchart Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 18)
cell.value = "ATS Flowchart S.D. of Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 19)
cell.value = "ATS Flowchart Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 20)
cell.value = "ATS Flowchart S.D. of Solution Time"
min_objective=999999999999999999
average_objective=0
average_time=0
Objective_Array=[]
Time_Array=[]
for i in range(num_of_run_ATSF):
    x_b,Time=ATS_Flowchart(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    average_time+=Time
    Objective_Array.append(x_b[0])
    Time_Array.append(Time)
average_time/=num_of_run_ATSF
average_objective/=num_of_run_ATSF
SD_Objective=0
SD_Time=0
for i in Objective_Array:
    SD_Objective+=pow(i-average_objective,2)
SD_Objective/=num_of_run_ATSF
SD_Objective=pow(SD_Objective,0.5)
for i in Time_Array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSF
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 16)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 17)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 18)
cell.value = SD_Objective
cell = Solo_Sheet.cell(row = 2, column = 19)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 20)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


num_of_run_ATST=1
cell = Solo_Sheet.cell(row = 1, column = 22)
cell.value = "ATS Text Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 23)
cell.value = "ATS Text Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 24)
cell.value = "ATS Text S.D. of Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 25)
cell.value = "ATS Text Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 26)
cell.value = "ATS Text S.D. of Solution Time"
min_objective=999999999999999999
average_objective=0
average_time=0
Objective_Array=[]
Time_Array=[]
for i in range(num_of_run_ATST):
    x_b,Time=ATS_Text(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    average_time+=Time
    Objective_Array.append(x_b[0])
    Time_Array.append(Time)
average_time/=num_of_run_ATST
average_objective/=num_of_run_ATST
SD_Objective=0
SD_Time=0
for i in Objective_Array:
    SD_Objective+=pow(i-average_objective,2)
SD_Objective/=num_of_run_ATST
SD_Objective=pow(SD_Objective,0.5)
for i in Time_Array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATST
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 22)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 23)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 24)
cell.value = SD_Objective
cell = Solo_Sheet.cell(row = 2, column = 25)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 26)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


num_of_run_HLSA=1
cell = Solo_Sheet.cell(row = 1, column = 28)
cell.value = "HLS Algorithm Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 29)
cell.value = "HLS Algorithm Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 30)
cell.value = "HLS Algorithm S.D. of Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 31)
cell.value = "HLS Algorithm Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 32)
cell.value = "HLS Algorithm S.D. of Solution Time"
min_objective=999999999999999999
average_objective=0
average_time=0
Objective_Array=[]
Time_Array=[]
for i in range(num_of_run_HLSA):
    x_b,Time=HLS_Algorithm(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    average_time+=Time
    Objective_Array.append(x_b[0])
    Time_Array.append(Time)
average_time/=num_of_run_HLSA
average_objective/=num_of_run_HLSA
SD_Objective=0
SD_Time=0
for i in Objective_Array:
    SD_Objective+=pow(i-average_objective,2)
SD_Objective/=num_of_run_HLSA
SD_Objective=pow(SD_Objective,0.5)
for i in Time_Array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_HLSA
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 28)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 29)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 30)
cell.value = SD_Objective
cell = Solo_Sheet.cell(row = 2, column = 31)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 32)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


num_of_run_HLSF=1
cell = Solo_Sheet.cell(row = 1, column = 34)
cell.value = "HLS Flowchart Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 35)
cell.value = "HLS Flowchart Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 36)
cell.value = "HLS Flowchart S.D. of Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 37)
cell.value = "HLS Flowchart Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 38)
cell.value = "HLS Flowchart S.D. of Solution Time"
min_objective=999999999999999999
average_objective=0
average_time=0
Objective_Array=[]
Time_Array=[]
for i in range(num_of_run_HLSF):
    x_b,Time=HLS_Flowchart(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    average_time+=Time
    Objective_Array.append(x_b[0])
    Time_Array.append(Time)
average_time/=num_of_run_HLSF
average_objective/=num_of_run_HLSF
SD_Objective=0
SD_Time=0
for i in Objective_Array:
    SD_Objective+=pow(i-average_objective,2)
SD_Objective/=num_of_run_HLSF
SD_Objective=pow(SD_Objective,0.5)
for i in Time_Array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_HLSF
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 34)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 35)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 36)
cell.value = SD_Objective
cell = Solo_Sheet.cell(row = 2, column = 37)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 38)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


num_of_run_HLST=1
cell = Solo_Sheet.cell(row = 1, column = 40)
cell.value = "HLS Text Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 41)
cell.value = "HLS Text Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 42)
cell.value = "HLS Text S.D. of Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 43)
cell.value = "HLS Text Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 44)
cell.value = "HLS Text S.D. of Solution Time"
min_objective=999999999999999999
average_objective=0
average_time=0
Objective_Array=[]
Time_Array=[]
for i in range(num_of_run_HLST):
    x_b,Time=HLS_Text(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    average_time+=Time
    Objective_Array.append(x_b[0])
    Time_Array.append(Time)
average_time/=num_of_run_HLST
average_objective/=num_of_run_HLST
SD_Objective=0
SD_Time=0
for i in Objective_Array:
    SD_Objective+=pow(i-average_objective,2)
SD_Objective/=num_of_run_HLST
SD_Objective=pow(SD_Objective,0.5)
for i in Time_Array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_HLST
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 40)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 41)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 42)
cell.value = SD_Objective
cell = Solo_Sheet.cell(row = 2, column = 43)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 44)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


# Number of RUNS of ATSAH for each Instance
num_of_run_ATSAH=Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 62)
cell.value = "ATSAH Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 63)
cell.value = "ATSAH Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 64)
cell.value = "ATSAH S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 65)
cell.value = "ATSAH Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 66)
cell.value = "ATSAH S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
for i in range(num_of_run_ATSAH):
    x_b,Time,Ultimate_counter=ATS_Algorithm_Heuristic(directory_details_for_saving=str(i))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
average_objective/=num_of_run_ATSAH
average_time/=num_of_run_ATSAH
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSAH
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSAH
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 62)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 63)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 64)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 65)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 66)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
"""


# Number of RUNS of ATSAML_6 for each Instance
num_of_run_ATSAML_6=0#Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 50)
cell.value = "ATSAML_6 Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 51)
cell.value = "ATSAML_6 Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 52)
cell.value = "ATSAML_6 S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 53)
cell.value = "ATSAML_6 Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 54)
cell.value = "ATSAML_6 S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
ATSAML6_Comparison_Sheet = Combined_Solution_Comparison.create_sheet("ATSAML6")
cell = ATSAML6_Comparison_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = ATSAML6_Comparison_Sheet.cell(row = 1, column = 2)
cell.value = "Objective Value"
cell = ATSAML6_Comparison_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
#for i in range(num_of_run_ATSAML_6):
while num_of_run_ATSAML_6<Heuristic_Runs:
    num_of_run_ATSAML_6+=1
    x_b,Time,Ultimate_counter,Powers=ATSAML_6(directory_details_for_saving=str(num_of_run_ATSAML_6))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
    cell = ATSAML6_Comparison_Sheet.cell(row = num_of_run_ATSAML_6+2, column = 1)
    cell.value = str(num_of_run_ATSAML_6)
    cell = ATSAML6_Comparison_Sheet.cell(row = num_of_run_ATSAML_6+2, column = 2)
    cell.value = str(x_b[0])
    cell = ATSAML6_Comparison_Sheet.cell(row = num_of_run_ATSAML_6+2, column = 3)
    cell.value = str(Time)
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
average_objective/=num_of_run_ATSAML_6
average_time/=num_of_run_ATSAML_6
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSAML_6
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSAML_6
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 50)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 51)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 52)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 53)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 54)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
out=cv2.VideoWriter('OutputVideo_ATSAML_6.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(1100,1000))
for i in range (1,num_of_run_ATSAML_6+1):
    imgA=cv2.imread(str(i)+"ATSAML_6 Evaluation/Progressive slowdown of Objective decreasePNG.png")
    imgB=cv2.imread(str(i)+"ATSAML_6 Evaluation/Understanding threshold effects of F_iterPNG.png")
    imgAB=cv2.vconcat([imgA,imgB])
    out.write(imgAB)
out.write(imgAB)
out.release()
out=cv2.VideoWriter('FullComparison_ATSAML_6.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,frameSize)
for i in range (1,num_of_run_ATSAML_6+1):
    imgC=cv2.imread(str(i)+"ATSAML_6 Evaluation/BoundsPNG.png")
    out.write(imgC)
out.write(imgC)
out.release()
out=cv2.VideoWriter('ParameterRadar_ATSAML_6.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(540,540))
for i in range (1,num_of_run_ATSAML_6+1):
    imgD=cv2.imread(str(i)+"ATSAML_6 Evaluation/ProgressParametersPNG.png")
    out.write(imgD)
out.write(imgD)
out.release()



# Number of RUNS of ATSAML_10 for each Instance
num_of_run_ATSAML_10=0#Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 56)
cell.value = "ATSAML_10 Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 57)
cell.value = "ATSAML_10 Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 58)
cell.value = "ATSAML_10 S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 59)
cell.value = "ATSAML_10 Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 60)
cell.value = "ATSAML_10 S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
ATSAML10_Comparison_Sheet = Combined_Solution_Comparison.create_sheet("ATSAML10")
cell = ATSAML10_Comparison_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = ATSAML10_Comparison_Sheet.cell(row = 1, column = 2)
cell.value = "Objective Value"
cell = ATSAML10_Comparison_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
#for i in range(num_of_run_ATSAML_10):
while num_of_run_ATSAML_10<Heuristic_Runs:
    num_of_run_ATSAML_10+=1
    x_b,Time,Ultimate_counter,Powers=ATSAML_10(directory_details_for_saving=str(num_of_run_ATSAML_10))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
    cell = ATSAML10_Comparison_Sheet.cell(row = num_of_run_ATSAML_10+2, column = 1)
    cell.value = str(num_of_run_ATSAML_10)
    cell = ATSAML10_Comparison_Sheet.cell(row = num_of_run_ATSAML_10+2, column = 2)
    cell.value = str(x_b[0])
    cell = ATSAML10_Comparison_Sheet.cell(row = num_of_run_ATSAML_10+2, column = 3)
    cell.value = str(Time)
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
average_objective/=num_of_run_ATSAML_10
average_time/=num_of_run_ATSAML_10
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSAML_10
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSAML_10
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 56)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 57)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 58)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 59)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 60)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
out=cv2.VideoWriter('OutputVideo_ATSAML_10.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(1100,1000))
for i in range (1,num_of_run_ATSAML_10+1):
    imgA=cv2.imread(str(i)+"ATSAML_10 Evaluation/Progressive slowdown of Objective decreasePNG.png")
    imgB=cv2.imread(str(i)+"ATSAML_10 Evaluation/Understanding threshold effects of F_iterPNG.png")
    imgAB=cv2.vconcat([imgA,imgB])
    out.write(imgAB)
out.write(imgAB)
out.release()
out=cv2.VideoWriter('FullComparison_ATSAML_10.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,frameSize)
for i in range (1,num_of_run_ATSAML_10+1):
    imgC=cv2.imread(str(i)+"ATSAML_10 Evaluation/BoundsPNG.png")
    out.write(imgC)
out.write(imgC)
out.release()
out=cv2.VideoWriter('ParameterRadar_ATSAML_10.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(540,540))
for i in range (1,num_of_run_ATSAML_10+1):
    imgD=cv2.imread(str(i)+"ATSAML_10 Evaluation/ProgressParametersPNG.png")
    out.write(imgD)
out.write(imgD)
out.release()




# Number of RUNS of ATSA_MT_19 for each Instance
num_of_run_ATSA_MT_19=0#Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 68)
cell.value = "ATSA_MT_19 Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 69)
cell.value = "ATSA_MT_19 Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 70)
cell.value = "ATSA_MT_19 S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 71)
cell.value = "ATSA_MT_19 Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 72)
cell.value = "ATSA_MT_19 S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
ATSAML19_Comparison_Sheet = Combined_Solution_Comparison.create_sheet("ATSAML19")
cell = ATSAML19_Comparison_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = ATSAML19_Comparison_Sheet.cell(row = 1, column = 2)
cell.value = "Objective Value"
cell = ATSAML19_Comparison_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
#for i in range(num_of_run_ATSA_MT_19):
while num_of_run_ATSA_MT_19<Heuristic_Runs:
    num_of_run_ATSA_MT_19+=1
    x_b,Time,Ultimate_counter,Powers=ATSA_MT_19(directory_details_for_saving=str(num_of_run_ATSA_MT_19))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
    cell = ATSAML19_Comparison_Sheet.cell(row = num_of_run_ATSA_MT_19+2, column = 1)
    cell.value = str(num_of_run_ATSA_MT_19)
    cell = ATSAML19_Comparison_Sheet.cell(row = num_of_run_ATSA_MT_19+2, column = 2)
    cell.value = str(x_b[0])
    cell = ATSAML19_Comparison_Sheet.cell(row = num_of_run_ATSA_MT_19+2, column = 3)
    cell.value = str(Time)
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
average_objective/=num_of_run_ATSA_MT_19
average_time/=num_of_run_ATSA_MT_19
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSA_MT_19
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSA_MT_19
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 68)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 69)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 70)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 71)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 72)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
out=cv2.VideoWriter('OutputVideo_ATSA_MT_19.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(1100,1000))
for i in range (1,num_of_run_ATSA_MT_19+1):
    imgA=cv2.imread(str(i)+"ATSA_MT_19 Evaluation/Progressive slowdown of Objective decreasePNG.png")
    imgB=cv2.imread(str(i)+"ATSA_MT_19 Evaluation/Understanding threshold effects of F_iterPNG.png")
    imgAB=cv2.vconcat([imgA,imgB])
    out.write(imgAB)
out.write(imgAB)
out.release()
out=cv2.VideoWriter('FullComparison_ATSA_MT_19.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,frameSize)
for i in range (1,num_of_run_ATSA_MT_19+1):
    imgC=cv2.imread(str(i)+"ATSA_MT_19 Evaluation/BoundsPNG.png")
    out.write(imgC)
out.write(imgC)
out.release()
out=cv2.VideoWriter('ParameterRadar_ATSAMT_19.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(540,540))
for i in range (1,num_of_run_ATSA_MT_19+1):
    imgD=cv2.imread(str(i)+"ATSA_MT_19 Evaluation/ProgressParametersPNG.png")
    out.write(imgD)
out.write(imgD)
out.release()




# JellyFishing Starts here

# Number of RUNS of ATSA_MT_6_JellyFish for each Instance
num_of_run_ATSAML_6_JF=0#Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 74)
cell.value = "ATSAML_6_JF Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 75)
cell.value = "ATSAML_6_JF Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 76)
cell.value = "ATSAML_6_JF S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 77)
cell.value = "ATSAML_6_JF Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 78)
cell.value = "ATSAML_6_JF S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
ATSAML6JF_Comparison_Sheet = Combined_Solution_Comparison.create_sheet("ATSAML6JF")
cell = ATSAML6JF_Comparison_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = ATSAML6JF_Comparison_Sheet.cell(row = 1, column = 2)
cell.value = "Objective Value"
cell = ATSAML6JF_Comparison_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
#for i in range(num_of_run_ATSAML_6_JF):
while num_of_run_ATSAML_6_JF<Heuristic_Runs:
    num_of_run_ATSAML_6_JF+=1
    x_b,Time,Ultimate_counter,Powers=ATSAMT_6_JF(directory_details_for_saving=str(num_of_run_ATSAML_6_JF))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
    cell = ATSAML6JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_6_JF+2, column = 1)
    cell.value = str(num_of_run_ATSAML_6_JF)
    cell = ATSAML6JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_6_JF+2, column = 2)
    cell.value = str(x_b[0])
    cell = ATSAML6JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_6_JF+2, column = 3)
    cell.value = str(Time)
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
average_objective/=num_of_run_ATSAML_6_JF
average_time/=num_of_run_ATSAML_6_JF
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSAML_6_JF
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSAML_6_JF
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 74)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 75)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 76)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 77)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 78)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
out=cv2.VideoWriter('OutputVideo_ATSAML_6_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(1100,1000))
for i in range (1,num_of_run_ATSAML_6_JF+1):
    imgA=cv2.imread(str(i)+"ATSAML_6_JF Evaluation/Progressive slowdown of Objective decreasePNG.png")
    imgB=cv2.imread(str(i)+"ATSAML_6_JF Evaluation/Understanding threshold effects of F_iterPNG.png")
    imgAB=cv2.vconcat([imgA,imgB])
    out.write(imgAB)
out.write(imgAB)
out.release()
out=cv2.VideoWriter('FullComparison_ATSAML_6_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,frameSize)
for i in range (1,num_of_run_ATSAML_6_JF+1):
    imgC=cv2.imread(str(i)+"ATSAML_6_JF Evaluation/BoundsPNG.png")
    out.write(imgC)
out.write(imgC)
out.release()
out=cv2.VideoWriter('ParameterRadar_ATSAML_6_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(540,540))
for i in range (1,num_of_run_ATSAML_6_JF+1):
    imgD=cv2.imread(str(i)+"ATSAML_6_JF Evaluation/ProgressParametersPNG.png")
    out.write(imgD)
out.write(imgD)
out.release()



# Number of RUNS of ATSAML_10_JF for each Instance
num_of_run_ATSAML_10_JF=0#Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 80)
cell.value = "ATSAML_10_JF Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 81)
cell.value = "ATSAML_10_JF Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 82)
cell.value = "ATSAML_10_JF S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 83)
cell.value = "ATSAML_10_JF Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 84)
cell.value = "ATSAML_10_JF S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
ATSAML10JF_Comparison_Sheet = Combined_Solution_Comparison.create_sheet("ATSAML10JF")
cell = ATSAML10JF_Comparison_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = ATSAML10JF_Comparison_Sheet.cell(row = 1, column = 2)
cell.value = "Objective Value"
cell = ATSAML10JF_Comparison_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
#for i in range(num_of_run_ATSAML_10_JF):
while num_of_run_ATSAML_10_JF<Heuristic_Runs:
    num_of_run_ATSAML_10_JF+=1
    x_b,Time,Ultimate_counter,Powers=ATSAMT_10_JF(directory_details_for_saving=str(num_of_run_ATSAML_10_JF))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
    cell = ATSAML10JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_10_JF+2, column = 1)
    cell.value = str(num_of_run_ATSAML_10_JF)
    cell = ATSAML10JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_10_JF+2, column = 2)
    cell.value = str(x_b[0])
    cell = ATSAML10JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_10_JF+2, column = 3)
    cell.value = str(Time)
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
average_objective/=num_of_run_ATSAML_10_JF
average_time/=num_of_run_ATSAML_10_JF
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSAML_10_JF
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSAML_10_JF
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 80)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 81)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 82)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 83)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 84)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
out=cv2.VideoWriter('OutputVideo_ATSAML_10_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(1100,1000))
for i in range (1,num_of_run_ATSAML_10_JF+1):
    imgA=cv2.imread(str(i)+"ATSAML_10_JF Evaluation/Progressive slowdown of Objective decreasePNG.png")
    imgB=cv2.imread(str(i)+"ATSAML_10_JF Evaluation/Understanding threshold effects of F_iterPNG.png")
    imgAB=cv2.vconcat([imgA,imgB])
    out.write(imgAB)
out.write(imgAB)
out.release()
out=cv2.VideoWriter('FullComparison_ATSAML_10_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,frameSize)
for i in range (1,num_of_run_ATSAML_10_JF+1):
    imgC=cv2.imread(str(i)+"ATSAML_10_JF Evaluation/BoundsPNG.png")
    out.write(imgC)
out.write(imgC)
out.release()
out=cv2.VideoWriter('ParameterRadar_ATSAML_10_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(540,540))
for i in range (1,num_of_run_ATSAML_10_JF+1):
    imgD=cv2.imread(str(i)+"ATSAML_10_JF Evaluation/ProgressParametersPNG.png")
    out.write(imgD)
out.write(imgD)
out.release()


# Number of RUNS of ATSAML_19_JF for each Instance
num_of_run_ATSAML_19_JF=0#Heuristic_Runs
cell = Solo_Sheet.cell(row = 1, column = 86)
cell.value = "ATSAML_19_JF Min. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 87)
cell.value = "ATSAML_19_JF Avg. Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 88)
cell.value = "ATSAML_19_JF S.D. of Objective Values"
cell = Solo_Sheet.cell(row = 1, column = 89)
cell.value = "ATSAML_19_JF Avg. Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 90)
cell.value = "ATSAML_19_JF S.D. of Solution Times"
min_objective=999999999999999999
average_objective=0
objective_array=[]
average_time=0
Time_array=[]
ATSAML19JF_Comparison_Sheet = Combined_Solution_Comparison.create_sheet("ATSAML19JF")
cell = ATSAML19JF_Comparison_Sheet.cell(row = 1, column = 1)
cell.value = "Instance"
cell = ATSAML19JF_Comparison_Sheet.cell(row = 1, column = 2)
cell.value = "Objective Value"
cell = ATSAML19JF_Comparison_Sheet.cell(row = 1, column = 3)
cell.value = "Solution Time"
#for i in range(num_of_run_ATSAML_19_JF):
while num_of_run_ATSAML_19_JF<Heuristic_Runs:
    num_of_run_ATSAML_19_JF+=1
    x_b,Time,Ultimate_counter,Powers=ATSAMT_19_JF(directory_details_for_saving=str(num_of_run_ATSAML_19_JF))
    if x_b[0]<min_objective:
        min_objective=x_b[0]
    average_objective+=x_b[0]
    objective_array.append(x_b[0])
    average_time+=Time
    Time_array.append(Time)
    cell = ATSAML19JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_19_JF+2, column = 1)
    cell.value = str(num_of_run_ATSAML_19_JF)
    cell = ATSAML19JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_19_JF+2, column = 2)
    cell.value = str(x_b[0])
    cell = ATSAML19JF_Comparison_Sheet.cell(row = num_of_run_ATSAML_19_JF+2, column = 3)
    cell.value = str(Time)
    Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
average_objective/=num_of_run_ATSAML_19_JF
average_time/=num_of_run_ATSAML_19_JF
SD_objective=0
for i in objective_array:
    SD_objective+=pow(i-average_objective,2)
SD_objective/=num_of_run_ATSAML_19_JF
SD_objective=pow(SD_objective,0.5)
SD_Time=0
for i in Time_array:
    SD_Time+=pow(i-average_time,2)
SD_Time/=num_of_run_ATSAML_19_JF
SD_Time=pow(SD_Time,0.5)
cell = Solo_Sheet.cell(row = 2, column = 86)
cell.value = min_objective
cell = Solo_Sheet.cell(row = 2, column = 87)
cell.value = average_objective
cell = Solo_Sheet.cell(row = 2, column = 88)
cell.value = SD_objective
cell = Solo_Sheet.cell(row = 2, column = 89)
cell.value = average_time
cell = Solo_Sheet.cell(row = 2, column = 90)
cell.value = SD_Time
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
out=cv2.VideoWriter('OutputVideo_ATSAML_19_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(1100,1000))
for i in range (1,num_of_run_ATSAML_19_JF+1):
    imgA=cv2.imread(str(i)+"ATSAML_19_JF Evaluation/Progressive slowdown of Objective decreasePNG.png")
    imgB=cv2.imread(str(i)+"ATSAML_19_JF Evaluation/Understanding threshold effects of F_iterPNG.png")
    imgAB=cv2.vconcat([imgA,imgB])
    out.write(imgAB)
out.write(imgAB)
out.release()
out=cv2.VideoWriter('FullComparison_ATSAML_19_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,frameSize)
for i in range (1,num_of_run_ATSAML_19_JF+1):
    imgC=cv2.imread(str(i)+"ATSAML_19_JF Evaluation/BoundsPNG.png")
    out.write(imgC)
out.write(imgC)
out.release()
out=cv2.VideoWriter('ParameterRadar_ATSAML_19_JF.avi',cv2.VideoWriter_fourcc(*'DIVX'),1,(540,540))
for i in range (1,num_of_run_ATSAML_19_JF+1):
    imgD=cv2.imread(str(i)+"ATSAML_19_JF Evaluation/ProgressParametersPNG.png")
    out.write(imgD)
out.write(imgD)
out.release()



cell = Solo_Sheet.cell(row = 1, column = 92)
cell.value = "BarisKara Exact Formulation Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 93)
cell.value = "BarisKara Exact Formulation Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 94)
cell.value = "MIP Gap %-age (or Best Bound)"
Objective,VT_Max_Utilised_Cap,Num_of_Vehicles_Used,Time,Gap,Best_Bound=BarisKara_Gurobi_Algorithm()
cell = Solo_Sheet.cell(row = 2, column = 92)
if Objective==-1:
    cell.value = "No feasble solution found"
else:
    cell.value = Objective
cell = Solo_Sheet.cell(row = 2, column = 93)
cell.value = Time
cell = Solo_Sheet.cell(row = 2, column = 94)
if Objective==-1:
    cell.value = Best_Bound
else:
    cell.value = str(Gap*100)+" %"
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


cell = Solo_Sheet.cell(row = 1, column = 96)
cell.value = "SanBan without Redundants Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 97)
cell.value = "SanBan without Redundants Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 98)
cell.value = "MIP Gap %-age (or Best Bound)"
Objective,VT_Max_Utilised_Cap,Num_of_Vehicles_Used,Time,Gap,Best_Bound=SANBAN_withoutRedundant_Algorithm()
cell = Solo_Sheet.cell(row = 2, column = 96)
if Objective==-1:
    cell.value = "No feasble solution found"
else:
    cell.value = Objective
cell = Solo_Sheet.cell(row = 2, column = 97)
cell.value = Time
cell = Solo_Sheet.cell(row = 2, column = 98)
if Objective==-1:
    cell.value = Best_Bound
else:
    cell.value = str(Gap*100)+" %"
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


"""
cell = Solo_Sheet.cell(row = 1, column = 46)
cell.value = "SanBan Exact Formulation Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 47)
cell.value = "SanBan Exact Formulation Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 48)
cell.value = "MIP Gap %-age (or Best Bound)"
Objective,VT_Max_Utilised_Cap,Num_of_Vehicles_Used,Time,Gap,Best_Bound=SANBAN_Gurobi_Algorithm()
cell = Solo_Sheet.cell(row = 2, column = 46)
if Objective==-1:
    cell.value = "No feasble solution found"
else:
    cell.value = Objective
cell = Solo_Sheet.cell(row = 2, column = 47)
cell.value = Time
cell = Solo_Sheet.cell(row = 2, column = 48)
if Objective==-1:
    cell.value = Best_Bound
else:
    cell.value = str(Gap*100)+" %"
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


cell = Solo_Sheet.cell(row = 1, column = 6)
cell.value = "Avci Topaloglu Exact Formulation Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 7)
cell.value = "Avci Topaloglu Exact Formulation Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 8)
cell.value = "Optimality Gap %-age (or Best Bound)"
Objective,Time,Gap,Best_Bound=Avci_Exact_Algorithm()
cell = Solo_Sheet.cell(row = 2, column = 6)
if Objective==-1:
    cell.value = "No feasble solution found"
else:
    cell.value = Objective
cell = Solo_Sheet.cell(row = 2, column = 7)
cell.value = Time
cell = Solo_Sheet.cell(row = 2, column = 8)
if Objective==-1:
    cell.value = Best_Bound
else:
    cell.value = str(Gap*100)+" %"
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")
"""



cell = Solo_Sheet.cell(row = 1, column = 100)
cell.value = "SanBan FutureWorks Objective Value"
cell = Solo_Sheet.cell(row = 1, column = 101)
cell.value = "SanBan FutureWorks Solution Time"
cell = Solo_Sheet.cell(row = 1, column = 102)
cell.value = "MIP Gap %-age (or Best Bound)"
Objective,VT_Max_Utilised_Cap,Num_of_Vehicles_Used,Time,Gap,Best_Bound=SANBAN_FutureWork()
cell = Solo_Sheet.cell(row = 2, column = 100)
if Objective==-1:
    cell.value = "No feasble solution found"
else:
    cell.value = Objective
cell = Solo_Sheet.cell(row = 2, column = 101)
cell.value = Time
cell = Solo_Sheet.cell(row = 2, column = 102)
if Objective==-1:
    cell.value = Best_Bound
else:
    cell.value = str(Gap*100)+" %"
Combined_Solution_Comparison.save("Combined Solution Comparison.xlsx")


for i in range(3):
    winsound.Beep(425, 125)
    winsound.Beep(575, 175)