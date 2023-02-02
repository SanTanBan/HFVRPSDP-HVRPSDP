import time
import matplotlib.pylab as plt
from gurobipy import *
import winsound
import os
import openpyxl
import GeneRead

# Refer:- https://www.gurobi.com/documentation/9.5/refman/variables.html
# Gurobi has some Tolerances on the Binary and Integer Variables...
# For Tolerances:- https://www.gurobi.com/documentation/9.5/refman/intfeastol.html#parameter:IntFeasTol
tolerance=-1e-5 # This is the default value for Gurobi

def BarisKara_Gurobi_Algorithm(max_seconds_allowed_for_calculation=86400,directory_containing_Node_Specifications_file="",directory__containing_Vehicle_Specifications_file="",directory_containing_Distance_Matrices_file="",directory_details_for_saving=""):
    
    os.mkdir(directory_details_for_saving+"BarisKara Exact Formulation")
    directory_to_save_Gurobi_solution=directory_details_for_saving+"BarisKara Exact Formulation/"

    Depot_and_Relief_Centres,Latitude,Longitude,PickUp,Delivery=GeneRead.Reader.Relief_Centre_Reader(directory_of_Relief_Centre_Specifications_file=directory_containing_Node_Specifications_file)
    CC=GeneRead.Reader.Distance_Combinations_Reader(directory_containing_Distance_Locations_file=directory_containing_Distance_Matrices_file,directory_of_Vehicle_Types_considered_file=directory__containing_Vehicle_Specifications_file)
    Vehicle_Types,VN,VQ,VS,VC,Vehicle_Width=GeneRead.Reader.Vehicle_Type_Reader(directory_of_Vehicle_Type_file=directory__containing_Vehicle_Specifications_file)

    c={}
    for key in CC:
        c[key]=CC[key]*VS[key[2]]
        #print(VS[key[2]])
        #print(c[key]," == ",CC[key])
    del CC


    Relief_Centres=Depot_and_Relief_Centres.copy()
    del Relief_Centres[0]
    Delivery[0]=0
    PickUp[0]=0
    Vehicle_Type_Maximum_Utilised_Capacity={}
    Number_of_Vehicle_used_of_each_Type={}


    x, y, z, t = {}, {}, {}, {}
    # Set the problem
    mdl=Model("VRP")
    mdl.setParam('TimeLimit', max_seconds_allowed_for_calculation)

    # Binary Decision Variables
    # Iff Arc joining i & j is included within the solution for the Layer k
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                x[i, j, k] = mdl.addVar(vtype=GRB.BINARY, name="x%d,%d,%d" % (i, j, k))
                mdl.update()

    # Continuous Decision Variables
    for k in Vehicle_Types:
        y[k] = mdl.addVar(vtype=GRB.INTEGER,name="x%d" % (k))
        mdl.update()
    m = mdl.addVar(name="m")
    mdl.update()
    # https://www.gurobi.com/documentation/9.5/refman/py_model_addvar.html#pythonmethod:Model.addVar
    # https://www.gurobi.com/documentation/9.5/refman/py_model_addvars.html

    # Auxilliary Decision Variables
    # Amount of delivery load across Arc(i,j) done by a Vehicle in Layer k
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            z[i, j] = mdl.addVar(name="x%d,%d" % (i, j))
            mdl.update()
    # Amount of pickup load across Arc(i,j) done by a Vehicle in Layer k
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            t[i, j] = mdl.addVar(name="x%d,%d" % (i, j))
            mdl.update()


    # Set Objective Function (Eq 1)
    mdl.setObjective(quicksum(VC[k]*y[k] for k in Vehicle_Types) + quicksum(x[i,j,k]*c[i,j,k] for k in Vehicle_Types for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j),GRB.MINIMIZE)
    mdl.update()

    # Eq2
    mdl.addConstr(quicksum(x[0,j,k] for j in Relief_Centres for k in Vehicle_Types) <= m)
    mdl.update()
    # Eq3
    mdl.addConstr(quicksum(x[i,0,k] for i in Relief_Centres for k in Vehicle_Types) <= m)
    mdl.update()

    #Eq4
    mdl.addConstrs(quicksum(x[i,j,k] for k in Vehicle_Types for i in Depot_and_Relief_Centres if i!=j) ==1 for j in Relief_Centres)
    mdl.update()

    # Eq5
    mdl.addConstrs(quicksum(x[i,j,k] for j in Depot_and_Relief_Centres if i!=j) == quicksum(x[j,i,k] for j in Depot_and_Relief_Centres if i!=j) for k in Vehicle_Types for i in Relief_Centres)
    mdl.update()

    # Eq6
    mdl.addConstrs(z[i,j]+t[i,j] <= quicksum(VQ[k]*x[i,j,k] for k in Vehicle_Types) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)
    mdl.update()

    """Corrected Equation 7""" #Eq7
    mdl.addConstrs(quicksum(z[j,i] for j in Depot_and_Relief_Centres if i!=j) - quicksum(z[i,j] for j in Depot_and_Relief_Centres if i!=j) == Delivery[i] for i in Relief_Centres)
    mdl.update()

    #Eq8
    mdl.addConstrs(quicksum(t[i,j] for j in Depot_and_Relief_Centres if i!=j)- quicksum(t[j,i] for j in Depot_and_Relief_Centres if i!=j) == PickUp[i] for i in Relief_Centres)
    mdl.update()

    # Eq9A
    mdl.addConstrs(quicksum(Delivery[j]*x[i,j,k] for k in Vehicle_Types) <= z[i,j] for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)
    mdl.update()
    # Eq9B
    mdl.addConstrs(z[i,j] <= quicksum((VQ[k]-Delivery[i])*x[i,j,k] for k in Vehicle_Types) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)
    mdl.update()

    # Eq10A
    mdl.addConstrs(quicksum(PickUp[i]*x[i,j,k] for k in Vehicle_Types) <= t[i,j] for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)
    mdl.update()
    # Eq10B
    mdl.addConstrs(t[i,j] <= quicksum((VQ[k]-PickUp[j])*x[i,j,k] for k in Vehicle_Types) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if i!=j)
    mdl.update()

    #Eq11
    mdl.addConstrs(t[0,j] == 0 for j in Relief_Centres)
    mdl.update()

    # Eq12
    mdl.addConstrs(z[i,0] == 0 for i in Relief_Centres)
    mdl.update()

    # Eq13
    mdl.addConstr(quicksum(y[k] for k in Vehicle_Types) <= m)
    mdl.update()

    # Eq14
    mdl.addConstrs(y[k] <= VN[k] for k in Vehicle_Types)
    mdl.update()

    # Eq15
    mdl.addConstrs(quicksum(x[0,j,k] for j in Relief_Centres) == y[k] for k in Vehicle_Types)
    mdl.update()

    mdl.optimize()
    #end_time=time.time()

    winsound.Beep(555, 888) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
    #print("This is the status:- ", status)
    #print('Runtime is',mdl.Runtime)

    # Solve the Problem using default CBC
    #status=prob.solve(p.Gurobi_CBC_CMD(maxSeconds=max_seconds_allowed_for_calculation, msg=1, gapRel=0))
    """
    Gurobi_start_time=time.time()
    if max_seconds_allowed_for_calculation>0:
        status=prob.solve(p.Gurobi_CBC_CMD(timeLimit=max_seconds_allowed_for_calculation))
    else:
        status=prob.solve()
    Gurobi_end_time=time.time()
    """
    #v=len(Vehicle_Types)+len(Depot_and_Relief_Centres)
    #winsound.Beep(333+19*v, 777+11*v) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
    #print("This is the status:- ", p.LpStatus[prob.status])
    
    OptimalityGap=mdl.MIPGap
    print("Final MIP gap value: %f" % OptimalityGap)
    best_bound=mdl.ObjBound
    print("The best Bound found is:- ",best_bound)

    Solutions_Found=mdl.SolCount
    objec_val=-1
    if Solutions_Found:
        objec_val=mdl.getObjective().getValue() #mdl.objVal

        Depot_First_Node=Depot_and_Relief_Centres[0]
        # Plotting the Depot and Relief Centres
        for k in Vehicle_Types:
            plt.figure(figsize=(11,11))
            for i in Depot_and_Relief_Centres:
                if i==Depot_First_Node:
                    plt.scatter(Longitude[i],Latitude[i], c='r',marker='s')
                    plt.text(Longitude[i] + 0.33, Latitude[i] + 0.33, "Depot")
                else:
                    plt.scatter(Longitude[i], Latitude[i], c='black')
                    plt.text(Longitude[i] + 0.33, Latitude[i] + 0.33, i)
            plt.title('mVRPSDC Tours for Vehicles of Type '+str(k)+" on the corresponding layer "+str(k))
            plt.ylabel("Latitude")
            plt.xlabel("Longitude")

            #routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j and p.value(x[i,j,k])==1]
            #routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j if x[i,j,k].x==1]
            routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres  if i!=j if ((x[i,j,k].x<=1-tolerance) and (x[i,j,k].x>=1+tolerance))]
            

            max=0   # Finding the Maximum Utilised Vehicle Capacity for this Vehicle Type
            for i, j in routes:
                utilized_capacity=t[i,j].x+z[i,j].x
                if utilized_capacity>max:
                    max=utilized_capacity
            Vehicle_Type_Maximum_Utilised_Capacity[k]=max
            print("\n The maximum vehicle capacity utilised ever in any tour in layer ",k," is: ",max," out of the total available",VQ[k])


            # Drawing the optimal routes Layerwise
            Current_Node=Depot_First_Node
            colour_intervals=len(routes)
            for counter in range(colour_intervals):
                for i,j in routes:
                    if i==Current_Node:
                        plt.annotate('', xy=[Longitude[j], Latitude[j]], xytext=[Longitude[i], Latitude[i]], arrowprops=dict(arrowstyle="-|>", connectionstyle='arc3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
                        #plt.annotate('', xy=[Longitude[j], Latitude[j]], xytext=[Longitude[i], Latitude[i]], arrowprops=dict(arrowstyle="simple", connectionstyle='angle3', edgecolor=(counter/colour_intervals,1-(counter/colour_intervals),1)))
                        
                        #Edge_Notes="P="+str(y[i,j,k].x)+" ; D="+str(z[i,j,k].x)
                        #plt.text((Longitude[i]+Longitude[j])/2, (Latitude[i]+Latitude[j])/2, f'{Edge_Notes}',fontweight="bold")
                        Current_Node=j
                        break
                routes.remove((i,j))

            used_vehicles=0 # Finding the maximum number of vehicles being used
            for j in Relief_Centres:
                used_vehicles=x[0,j,k].x+used_vehicles
            if used_vehicles!=0:
                Number_of_Vehicle_used_of_each_Type[k]=used_vehicles
            print("\n The maximum numbers of vehicles used is: ",used_vehicles," out of total available ",VN[k])
            name="Vehicles_ "+str(used_vehicles)+"--"+str(VN[k])+" and Capacity_ "+str(max)+"--"+str(VQ[k])+".eps"
            main_dir_for_Image=directory_to_save_Gurobi_solution+"{}"
            plt.savefig(main_dir_for_Image.format(name),format='eps')

        # Writing the Routes in a Text File

        textfile = open(directory_to_save_Gurobi_solution+"Vehicle Routes as per Gurobi.txt","w")
        if max_seconds_allowed_for_calculation>0:
            textfile.write("\t The problem was stopped before the given time-limit of "+str(max_seconds_allowed_for_calculation)+".\n")
            textfile.write("\t The Objective Value is "+str(objec_val)+" obtained within "+str(mdl.Runtime)+" seconds with Relative Gap of "+str(OptimalityGap*100)+"% \n \n \n")
        else:
            #textfile.write("\t The Status of the problem is "+p.LpStatus[prob.status]+" \n")
            textfile.write("\t The Objective Value is "+str(objec_val)+" obtained within "+str(mdl.Runtime)+" seconds \n \n \n")
        for k in Vehicle_Types:
            counter=0
            for j in Relief_Centres:
                #if x[0,j,k].x==1:
                if ((x[0,j,k].x<=1-tolerance) and (x[0,j,k].x>=1+tolerance)):
                    counter+=1
                    start_node=j
                    textfile.write("Vehicle Type: "+str(k)+",\t Vehicle Number: "+str(counter)+", \t Route=\t 0")
                    while start_node!=0:
                        textfile.write(" --> "+str(start_node))
                        for i in Depot_and_Relief_Centres:
                            #if  start_node!=i and x[start_node,i,k].x==1:
                            if  start_node!=i and ((x[start_node,i,k].x<=1-tolerance) and (x[start_node,i,k].x>=1+tolerance)):
                                start_node=i
                                break
                    if start_node==0:
                        textfile.write(" --> "+str(start_node)+"\n")
        for k in Vehicle_Type_Maximum_Utilised_Capacity:
            if Vehicle_Type_Maximum_Utilised_Capacity[k]>0:
                textfile.write("\n Vehicle Type "+str(k)+": Maximum Utilised Capacity by any Vehicle = "+str(Vehicle_Type_Maximum_Utilised_Capacity[k])+" units out of the total available "+str(VQ[k])+",\t and Number of Vehicles used is "+str(Number_of_Vehicle_used_of_each_Type[k])+" out of Total allowed "+str(VN[k]))
                #textfile.write("\n Maximum Vehicle Capacity Utilised by any Vehicle of Type "+str(k)+" is "+Vehicle_Type_Maximum_Utilised_Capacity[k]+" units out of the total available "+VQ[k]+"\n")
        textfile.close()


        # Call a Workbook() function of openpyxl to create a new blank Workbook object
        wb_individual = openpyxl.Workbook()
        # Get workbook active sheet from the active attribute
        sheet_individual = wb_individual.active
        row_number_on_Individual_Sheet=1
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
        cell.value = "From Node i"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
        cell.value = "To Node j"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
        cell.value = "Vehicle Type k"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
        cell.value = "x_ijk indicating whether the Arc is selected"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
        cell.value = "t_ij indicating the amount of Pickup"
        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
        cell.value = "z_ij indicating the amount of Delivery"
        for i in Depot_and_Relief_Centres:
            for j in Depot_and_Relief_Centres:
                for k in Vehicle_Types:
                    if i!=j:
                        row_number_on_Individual_Sheet+=1
                        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 1)
                        cell.value = i
                        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 2)
                        cell.value = j
                        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 3)
                        cell.value = k
                        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 4)
                        cell.value = x[i,j,k].x
                        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 5)
                        cell.value = t[i,j].x
                        cell = sheet_individual.cell(row = row_number_on_Individual_Sheet, column = 6)
                        cell.value = z[i,j].x
        wb_individual.save(str(directory_to_save_Gurobi_solution)+"Solution Details.xlsx")

    return objec_val,Vehicle_Type_Maximum_Utilised_Capacity,Number_of_Vehicle_used_of_each_Type,mdl.Runtime,OptimalityGap,best_bound


# For retrieving detailed Solutions
# https://www.gurobi.com/documentation/9.5/refman/retrieving_solutions.html
# https://www.gurobi.com/documentation/9.5/refman/objbound.html#attr:ObjBound
# https://www.gurobi.com/documentation/9.5/refman/parameter_examples.html#sec:ParameterExamples